'''
 fmt = getattr(settings, 'LOG_FORMAT', None)
 lvl = getattr(settings, 'LOG_LEVEL', logging.DEBUG)

 logging.basicConfig(format=fmt, level=lvl)
 logging.debug(neringAmarbject.OrderId.id)
 '''
from django.shortcuts import render
from django.shortcuts import get_object_or_404
from django.http import JsonResponse
from django.template.loader import render_to_string
from django.db.models import Sum
import jdatetime
from django.http import HttpResponseRedirect
from django.http import HttpResponse
from django.views.decorators import csrf
import django.core.serializers
import logging
from django.conf import settings

from cmms.models import TolidAmar,TolidMoshakhase
#from django.core import serializers
import json
from django.forms.models import model_to_dict
from cmms.forms import RingAmarForm
from django.urls import reverse_lazy
from django.db import transaction
from django.contrib.auth.decorators import permission_required
from django.contrib.auth.context_processors import PermWrapper
from rest_framework.decorators import api_view
from cmms.api.WOSerializer import *
from rest_framework.response import Response
from rest_framework import status
from cmms.business.amarutility import AmarUtility
from cmms.business.DateJob import *
from django.db.models import Max
from django.db import IntegrityError
from django.views.decorators.csrf import csrf_exempt
import openpyxl
# from openpyxl import Workbook
# from openpyxl.utils.dataframe import dataframe_to_rows
# import pandas as pd
@permission_required('cmms.view_ringamar')
def list_tolidAmar(request,id=None):
    #
    books = TolidAmar.objects.none()
    # wo=AmarUtility.doPaging(request,books)

    assetMakan=Asset.objects.filter(assetTypes=1,assetIsLocatedAt__isnull=True)
    return render(request, 'cmms/tolidamar/tolidAmarList.html', {'ringAmar': books,'section':'list_tolidAmar','makan':assetMakan})



def getAmarTolidMoshakhase(request):
    data=dict()
    qry=request.GET.get("qry",False)
    if(qry):
        results=TolidMoshakhase.objects.filter(mogheiat=qry)
        if(results.count()>0):
            data["result"]={'mogheiat':results[0].mogheiat,'keifiat':results[0].keyfiat,'id':results[0].id,'vaziat':results[0].vaziat}
        else:
            data["result"]={}
        return JsonResponse(data, safe=False)
    return data

def loadAmarTolidTableInfo(request):
    makan=request.GET.get("makan",False)
    dt=request.GET.get("dt",False)
    date1=DateJob.getTaskDate(dt)
    data=dict()
    x1=''
    y1=''
    if(makan):
        amar=TolidAmar.objects.filter(location=makan,registered_date=date1).order_by('registered_date')
        location=Asset.objects.get(id=makan)

        if(amar.count()==0):
            amar=[]


            j=TolidAmar.objects.create(location=location,registered_date=date1,tedad=0,meghdar=0)
            amar.append(j)
                # data['amar']= render_to_string('cmms/tolidamar/partialRingAmarList2.html', {
                # 'dt':dt,
                # 'date':date1,
                # 'makan':makan,
                # 'makan_id':location.id,
                # 'perms': PermWrapper(request.user)})

            data['amar']= render_to_string('cmms/tolidamar/partialRingAmarList2.html', {
            'tolidAmar': amar,
            'perms': PermWrapper(request.user),

        })
        data['amar']= render_to_string('cmms/tolidamar/partialRingAmarList2.html', {
        'tolidAmar': amar,
        'perms': PermWrapper(request.user)
        })

        data['form_is_valid']=True

    return JsonResponse(data)
@csrf_exempt
def saveAmarTolidTableInfo(request):
    # print(request.body)
    # print(request.POST)
    data = json.loads(request.body)
    dt={}
    print("********")
    for i in data:
        # print("********")


        if('id' in i):

            moshakhase=0
            if(int(i['id'])!=-1):
                print(i['radif'],i['tolidmoshakhase'])
                if(int(i['tolidmoshakhase'])==0):
                        print("here!",i['vaziat_text'])
                        nakh_obj=TolidMoshakhase.objects.create(mogheiat=i['vaziat_text'],keyfiat=i['keyfiat'],vaziat=i['mogheiat'])
                        moshakhase=nakh_obj.id
                        print(nakh_obj)
                else:
                    moshakhase=i['tolidmoshakhase']


            # -2 show new row and -1 represent header row which is not contain ant information
            if(i['id']!='-2'):

                if(int(i['id'])!=-1):
                    print(i['isheatset'])
                    amar=TolidAmar.objects.get(id=i['id'])
                    amar.tolidmoshakhase=TolidMoshakhase.objects.get(id=moshakhase)
                    amar.registered_date=i["registered_date"].replace('/','-')
                    amar.tedad=float(i["tedad"])
                    amar.meghdar=float(i["meghdar"])
                    amar.location=Asset.objects.get(id=i["location"])
                    amar.isheatset=i['isheatset']
                    amar.save()

            #-2 show new row
            if(i['id']=='-2'):
                print(moshakhase)
                isheatset=i['isheatset']
                newRow=TolidAmar.objects.create(tolidmoshakhase=TolidMoshakhase.objects.get(id=moshakhase),registered_date=i["registered_date"].replace('/','-'),tedad=int(i["tedad"]),meghdar=float(i["meghdar"]),location=Asset.objects.get(id=i["location"]),isheatset=isheatset)
                dt[i['radif']]=newRow.id
    return JsonResponse(dt)

def col_letter_to_index(letter):
    index = 0
    for char in letter:
        index = index * 26 + (ord(char.upper()) - ord('A')) + 1
    return index - 1  # Subtract 1 to make it 0-based index
def tolidImport(request):
    return render(request, 'cmms/tolidamar/amarUpload.html', {'location':request.GET.get('location',False)})
def upload_file_tolidamar(request):
    def iter_rows(ws):
        for row in ws.iter_rows():
            yield [cell.value for cell in row]
    if request.method == 'POST':
        my_file=request.FILES.get('file')
        date_pattern = r'^\d{4}-\d{2}-\d{2}$'
# Replace 'your_excel_file.xlsx' with the path to your Excel file
        workbook = openpyxl.load_workbook(my_file)
        # Specify the sheet name
        sheet = workbook['Sheet1']  # Replace 'Sheet1' with your sheet name
        i=1
        print(request.GET.get("location",False))

        # Or use the default sheet (usually the first one)
        sheet = workbook.active
        for row in sheet.iter_rows(values_only=True):
            if(row[col_letter_to_index('x')] is not None and row[col_letter_to_index('v')] is not None):


                nomre_nakh=TolidMoshakhase.objects.filter(mogheiat=row[col_letter_to_index('ac')])
                tarikh=row[col_letter_to_index('x')].replace('/','-')
                if(nomre_nakh.count()>0):
                    print("exist")
                    location=Asset.objects.get(id=int(request.GET.get("location",False)))
                    registered_date=DateJob.getTaskDate(tarikh)
                    tedad=row[col_letter_to_index('r')]
                    meghdar=row[col_letter_to_index('p')]
                    print(i,row[col_letter_to_index('p')])
                    isheatset=False if 'HB' in nomre_nakh[0].vaziat else True
                    TolidAmar.objects.create(location=location,registered_date=registered_date,tedad=tedad,meghdar=meghdar,tolidmoshakhase=nomre_nakh[0])
                else:
                    print("new")
                    mogheiat=row[col_letter_to_index('ac')]
                    keyfiat=row[col_letter_to_index('v')]
                    vaziat=row[col_letter_to_index('z')]
                    nomre_nakh=TolidMoshakhase.objects.create(vaziat=vaziat,mogheiat=mogheiat,keyfiat=keyfiat)
                    location=Asset.objects.get(id=int(request.GET.get("location",False)))
                    registered_date=DateJob.getTaskDate(tarikh)
                    tedad=row[col_letter_to_index('r')]
                    meghdar=row[col_letter_to_index('p')]
                    isheatset=False if 'HB' in nomre_nakh.vaziat else True
                    TolidAmar.objects.create(location=location,registered_date=registered_date,tedad=tedad,meghdar=meghdar,tolidmoshakhase=nomre_nakh)
                i=i+1
                print(i)






                    # TolidAmar.objects.create


            # Do something with the cell values

        data=dict()

        return JsonResponse(data)
    return JsonResponse({'post':'fasle'})
